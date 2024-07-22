import os
import datetime
import pandas as pd
import openpyxl
from tkinter import Tk, Label, Button, filedialog, StringVar
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import threading
import sys

def read_data(file_path):
    data = pd.read_excel(file_path, sheet_name=None)
    return data

def copy_styles(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                if cell.fill.fill_type:
                    target_cell.fill = PatternFill(start_color=cell.fill.start_color,
                                                   end_color=cell.fill.end_color,
                                                   fill_type=cell.fill.fill_type)
                if cell.font:
                    target_cell.font = Font(name=cell.font.name,
                                            size=cell.font.size,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            underline=cell.font.underline,
                                            strike=cell.font.strike,
                                            color=cell.font.color)
                if cell.alignment:
                    target_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                      vertical=cell.alignment.vertical,
                                                      text_rotation=cell.alignment.text_rotation,
                                                      wrap_text=cell.alignment.wrap_text,
                                                      shrink_to_fit=cell.alignment.shrink_to_fit,
                                                      indent=cell.alignment.indent)
    for merge_cell in source_sheet.merged_cells:
        target_sheet.merge_cells(str(merge_cell))

def adjust_column_width(sheet):
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                text = str(cell.value)
                if isinstance(cell.value, datetime.datetime):
                    text = cell.value.strftime('%Y-%m-%d %H:%M:%S')
                    cell_length = len(text) + 4
                elif isinstance(cell.value, str):
                    cell_length = sum(2 if ord(char) > 256 else 1 for char in text) + 1
                else:
                    cell_length = len(text)

                column = cell.column_letter
                column_widths[column] = max(column_widths.get(column, 0), cell_length)

    for column, max_width in column_widths.items():
        sheet.column_dimensions[column].width = max_width if max_width < 100 else 100

def process_excel(file_path, output_path, callback):
    data = read_data(file_path)
    new_workbook = openpyxl.Workbook()
    first_sheet = True

    for sheet_name, df in data.items():
        if first_sheet:
            new_sheet = new_workbook.active
            new_sheet.title = sheet_name
            first_sheet = False
        else:
            new_sheet = new_workbook.create_sheet(title=sheet_name)

        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)

        original_wb = openpyxl.load_workbook(file_path, data_only=True)
        original_sheet = original_wb[sheet_name]

        copy_styles(original_sheet, new_sheet)
        original_wb.close()

        adjust_column_width(new_sheet)

    new_workbook.save(output_path)
    callback(f"已处理完成，文件已保存到 {output_path}")

def start_thread(file_path, status_label, root):
    if file_path:
        output_path = os.path.join(os.path.dirname(file_path), "修改后" + os.path.basename(file_path))
        status_label.set("处理中，请等待...")
        threading.Thread(target=process_excel, args=(file_path, output_path, lambda message: root.after(0, status_label.set, message))).start()
    else:
        status_label.set("没有选择文件")

def create_gui():
    root = Tk()
    root.title("Excel文件处理器")
    root.geometry("500x150")

    status_label = StringVar()
    status_label.set("请选择一个文件进行处理！\n 处理时间因复杂程度而异！请耐心等待，不要重复点击！")

    def open_file():
        file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])
        start_thread(file_path, status_label, root)

    Label(root, textvariable=status_label).pack()
    Button(root, text="选择文件并处理", command=open_file).pack()

    root.mainloop()

if __name__ == "__main__":
    create_gui()
