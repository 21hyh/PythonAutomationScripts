import datetime
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def read_data(file_path):
    # 使用pandas读取数据
    data = pd.read_excel(file_path, sheet_name=None)  # 读取所有工作表
    return data


def copy_styles(source_sheet, target_sheet):
    # 复制样式，主要是颜色、合并单元格、字体和对齐方式
    for row in source_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:  # 只处理非空单元格
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                # 复制颜色
                if cell.fill.fill_type:
                    target_cell.fill = PatternFill(start_color=cell.fill.start_color,
                                                   end_color=cell.fill.end_color,
                                                   fill_type=cell.fill.fill_type)
                # 复制字体样式
                if cell.font:
                    target_cell.font = Font(name=cell.font.name,
                                            size=cell.font.size,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            underline=cell.font.underline,
                                            strike=cell.font.strike,
                                            color=cell.font.color)
                # 复制对齐方式
                if cell.alignment:
                    target_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                      vertical=cell.alignment.vertical,
                                                      text_rotation=cell.alignment.text_rotation,
                                                      wrap_text=cell.alignment.wrap_text,
                                                      shrink_to_fit=cell.alignment.shrink_to_fit,
                                                      indent=cell.alignment.indent)
    # 复制合并单元格
    for merge_cell in source_sheet.merged_cells:
        target_sheet.merge_cells(str(merge_cell))




def adjust_column_width(sheet):
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                # 将单元格内容转换为字符串
                text = str(cell.value)
                if isinstance(cell.value, datetime.datetime):  # 检查是否为日期类型
                    # 采用特定的日期显示格式
                    text = cell.value.strftime('%Y-%m-%d %H:%M:%S')
                    # 对于日期，设置一个较宽的固定宽度，或者按格式化后的长度计算
                    cell_length = len(text)+4  # 添加缓冲
                elif isinstance(cell.value, str):
                    # 对于字符串，特别是包含中文的情况，给予更多的宽度
                    cell_length = sum(2 if ord(char) > 256 else 1 for char in text) + 1  # 添加缓冲
                else:
                    # 对于数字和其他类型，使用它们的字符串长度
                    cell_length = len(text)

                # 获取列标识符
                column = cell.column_letter
                # 更新该列的最大宽度
                column_widths[column] = max(column_widths.get(column, 0), cell_length)

    # 应用计算出的列宽，每列宽度加上额外的2个单位的缓冲
    for column, max_width in column_widths.items():
        adjusted_width = max_width
        sheet.column_dimensions[column].width = adjusted_width if adjusted_width < 100 else 100  # 避免异常宽度



def process_excel(file_path, output_path):
    # 读取数据
    data = read_data(file_path)

    # 使用openpyxl创建新的Excel文件
    new_workbook = openpyxl.Workbook()
    first_sheet = True

    for sheet_name, df in data.items():
        if first_sheet:
            new_sheet = new_workbook.active
            new_sheet.title = sheet_name
            first_sheet = False
        else:
            new_sheet = new_workbook.create_sheet(title=sheet_name)

        # 使用pandas的数据填充新工作表
        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)

        # 使用非只读模式加载原始工作簿以复制样式
        original_wb = openpyxl.load_workbook(file_path, data_only=True)
        original_sheet = original_wb[sheet_name]

        # 复制样式到新工作表
        copy_styles(original_sheet, new_sheet)

        # 关闭原工作簿以释放资源
        original_wb.close()

        # 调整新工作表的列宽
        adjust_column_width(new_sheet)

    # 保存新的Excel文件
    new_workbook.save(output_path)


# 调用函数处理文件
process_excel('D:/30866/Desktop/a1.xlsx', 'D:/30866/Desktop/a1_processed4.xlsx')